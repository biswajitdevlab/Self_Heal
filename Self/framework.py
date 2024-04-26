import xlsxwriter
import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time

class SelfHealingAutomationFramework:
    def __init__(self, url=None):
        self.driver = webdriver.Chrome()
        self.max_retries = 3
        self.url = url

    def find_element_with_retry(self, locators):
        retries = 0
        while retries < self.max_retries:
            for by, value in locators:
                try:
                    element = self.driver.find_element(by, value)
                    return element
                except NoSuchElementException:
                    continue
            retries += 1
            print(f"Element not found, retrying... (Attempt {retries}/{self.max_retries})")
            time.sleep(2)  # Wait before retrying
        raise NoSuchElementException(f"Element not found after {self.max_retries} retries")

    def extract_locators_to_excel(self, url):
        """
        Extract locators from the web page and save them to an Excel file.
        """
        self.driver.get(url)
        elements = self.driver.find_elements(By.XPATH, '//*')

        locators = []
        for element in elements:
            xpath_locator = self.generate_xpath_locator(element)
            css_locator = self.generate_css_locator(element)
            locators.append((xpath_locator, css_locator))

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Locators"
        sheet.append(["XPath", "CSS Selector"])
        for xpath_locator, css_locator in locators:
            sheet.append([xpath_locator, css_locator])

        try:
            wb.save("locators.xlsx")
            print("Locators extracted and saved to locators.xlsx")
        except Exception as e:
            print(f"Error occurred while saving locators: {e}")

    def generate_xpath_locator(self, element):
        tag_name = element.tag_name
        xpath_locator = ".//" + tag_name
        parent = element
        while parent.tag_name != "html":
            parent = parent.find_element(By.XPATH, "..")  # Change this line
            sibling_index = len(parent.find_elements(By.XPATH, f"./{tag_name}[preceding-sibling::{tag_name}]")) + 1
            xpath_locator = f"./{tag_name}[{sibling_index}]/" + xpath_locator
        return xpath_locator

    def generate_css_locator(self, element):
        tag_name = element.tag_name
        css_locator = tag_name
        parent = element
        while parent.tag_name != "html":
            parent = parent.find_element(By.XPATH, "..")  # Change this line
            sibling_index = len(parent.find_elements(By.CSS_SELECTOR,
                                                     f"{tag_name}:nth-child(-n+{len(parent.find_elements(By.XPATH, f'./{tag_name}')) + 1})")) + 1
            css_locator = f"{tag_name}:nth-child({sibling_index}) > " + css_locator
        return css_locator

    def read_locators_from_excel(self, excel_file):
        # Implementation for reading locators from Excel file
        locators = []
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            locators.append((By.XPATH, row[0]))
            locators.append((By.CSS_SELECTOR, row[1]))
        return locators

    def run_test_case(self, test_case, locators):
        for step in test_case:
            action = step['action']
            if action == 'click':
                element_locators = locators.get(step['element'])
                element = self.find_element_with_retry(element_locators)
                element.click()
            elif action == 'input_text':
                element_locators = locators.get(step['element'])
                element = self.find_element_with_retry(element_locators)
                element.send_keys(step['text'])

    def run(self, test_cases):
        locators = self.read_locators_from_excel("locators.xlsx")
        for test_case in test_cases:
            print(f"Running test case: {test_case['name']}")
            try:
                self.run_test_case(test_case['steps'], locators)
                print("Test case passed.")
            except Exception as e:
                print(f"Test case failed: {str(e)}")
                print("Attempting self-healing...")
                if self.url:  # Ensure that the URL is provided
                    self.extract_locators_to_excel(self.url)  # Extract new locators
                    locators = self.read_locators_from_excel("locators.xlsx")  # Reload locators
                    try:
                        self.run_test_case(test_case['steps'], locators)  # Retry test case
                        print("Test case passed after self-healing.")
                    except Exception as e:
                        print(f"Test case failed after self-healing: {str(e)}")
                else:
                    print("URL not provided. Skipping self-healing.")
            finally:
                # Clean up or reset state as needed
                pass
        self.driver.quit()